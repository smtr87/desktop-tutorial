

#this code is successfully tested in Google Colab (https://colab.research.google.com/)
#https://colab.research.google.com/drive/1A5WhgvK3KmkoJhCVgAaRayErVFGyPZ7A?usp=sharing
#get "test.xlsx' file from here https://docs.google.com/spreadsheets/d/18N-KMXrjR6anVWnSMItEtgzU42NNm5sj/edit?usp=sharing&ouid=109735252067534170443&rtpof=true&sd=true
import openpyxl
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import sys
sys.path.insert(0,'/usr/lib/chromium-browser/chromedriver')

import time
import pandas as pd
#from bs4 import BeautifulSoup
from selenium import webdriver
import chromedriver_autoinstaller

# setup chrome optionspui
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless') # ensure GUI is off
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# set path to chromedriver as per your configuration
chromedriver_autoinstaller.install()

# set the target URL
url = "put-url-here-to-scrape"

wb= openpyxl.load_workbook('/workspaces/desktop-tutorial/test.xlsx')
#getting weekday as dayname. i.e sunday ,monday
week_days=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
day=week_days[date.today().weekday()]
print('Today is:',day)
#ws=wb.active
#seelcting sheet from the workbook based on the current day
wb_sheet= wb[day]
#defining webdrive (running web browser)
driver = webdriver.Chrome(options=chrome_options)
# without this sometime it returns no data because of net delay
driver.implicitly_wait(10)
for i in range(3,13,1):
  driver = webdriver.Chrome(options=chrome_options)
  #take content from excel file
  scell=wb_sheet.cell(row=i,column=3)
  search=scell.value
  #automation
  driver.get('https://www.google.com')
  search_box = driver.find_element(By.NAME,'q')
  search_box.send_keys(search)
  suggestion_list= driver.find_elements(By.XPATH,'//*[@id="Alh6id"]/div[1]/div/ul/li')
  #to test search automation is getting data
  # for x in suggestion_list:
  #  print(x.text)

  #to find longest search suggestion
  w_max_cell=wb_sheet.cell(row=i,column=4)
  max=0
  for y in suggestion_list:
    le=int(len(y.text))
    if (le>max):
      max=le
      w_max_cell.value=y.text

  #to find shortest search suggestion
  w_min_cell=wb_sheet.cell(row=i,column=5)
  min=max
  for z in suggestion_list:
    le=int(len(z.text))
    if (le<min):
      min=le
      w_min_cell.value=z.text

  #save to excel file
  wb.save('/workspaces/desktop-tutorial/test.xlsx')
  #to check writing is successful
  write1=w_max_cell.value
  write2=w_min_cell.value
  print('\n','>>>LONGEST OPTION:',write1)
  print('>>>SHORTEST OPTION:',write2,'\n')
  driver.quit()
